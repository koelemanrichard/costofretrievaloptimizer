from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# Define styles
header_fill = PatternFill('solid', fgColor='1F4E79')
header_font = Font(bold=True, color='FFFFFF', size=11)
section_fill = PatternFill('solid', fgColor='BDD7EE')
critical_fill = PatternFill('solid', fgColor='FF6B6B')
critical_font = Font(bold=True, color='FFFFFF')
high_fill = PatternFill('solid', fgColor='FFA500')
medium_fill = PatternFill('solid', fgColor='FFD93D')
success_fill = PatternFill('solid', fgColor='6BCB77')
warning_fill = PatternFill('solid', fgColor='FFE66D')
gray_fill = PatternFill('solid', fgColor='E0E0E0')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(sheet, row, cols):
    for col in range(1, cols + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def style_data_cell(sheet, row, col, align='left'):
    cell = sheet.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = left_align if align == 'left' else center_align

# ============ SHEET 1: Pre-Migration Checklist ============
ws1 = wb.active
ws1.title = 'Pre-Migration Checklist'

ws1['A1'] = 'MVGM VvE Beheer - Pre-Migration Checklist'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws1.merge_cells('A1:F1')

ws1['A2'] = 'Migration: mvgm.com to mvgm-vvebeheer.nl'
ws1['A2'].font = Font(italic=True, size=10)
ws1.merge_cells('A2:F2')

headers1 = ['Category', 'Task', 'Owner', 'Due Date', 'Status', 'Notes']
for col, header in enumerate(headers1, 1):
    ws1.cell(row=4, column=col, value=header)
style_header_row(ws1, 4, 6)

checklist_data = [
    ['Benchmark', 'Export current keyword rankings for VvE terms', 'SEO Team', 'Before Phase 1', 'Not Started', 'Focus on vve beheer + city terms'],
    ['Benchmark', 'Screenshot Google Analytics traffic data (30 days)', 'SEO Team', 'Before Phase 1', 'Not Started', 'Document organic sessions'],
    ['Benchmark', 'Run site:mvgm.com/nl/vastgoeddiensten/vve-beheer', 'SEO Team', 'Before Phase 1', 'Not Started', 'Count indexed pages'],
    ['Benchmark', 'Save GSC backlink data export', 'SEO Team', 'Before Phase 1', 'Completed', 'Done - 114 backlinks identified'],
    ['New Site', 'Verify all destination URLs return 200', 'Dev Team', 'Before Phase 1', 'Not Started', 'Test all 113 target URLs'],
    ['New Site', 'Confirm canonical tags point to new domain', 'Dev Team', 'Before Phase 1', 'Not Started', 'No canonicals to mvgm.com'],
    ['New Site', 'Check no pages have noindex', 'Dev Team', 'Before Phase 1', 'Not Started', 'All VvE pages must be indexable'],
    ['New Site', 'Create XML sitemap', 'Dev Team', 'Before Phase 1', 'Not Started', 'mvgm-vvebeheer.nl/sitemap.xml'],
    ['New Site', 'Verify robots.txt allows crawling', 'Dev Team', 'Before Phase 1', 'Not Started', ''],
    ['New Site', 'Confirm HTTPS/SSL certificate working', 'Dev Team', 'Before Phase 1', 'Not Started', ''],
    ['New Site', 'Test mobile-friendliness', 'Dev Team', 'Before Phase 1', 'Not Started', 'Google Mobile-Friendly Test'],
    ['GSC Setup', 'Add mvgm-vvebeheer.nl as new property', 'SEO Team', 'Before Phase 1', 'Not Started', ''],
    ['GSC Setup', 'Verify ownership (DNS or HTML)', 'SEO Team', 'Before Phase 1', 'Not Started', ''],
    ['GSC Setup', 'Submit sitemap to GSC', 'SEO Team', 'Before Phase 1', 'Not Started', ''],
    ['Redirects', 'Test redirect rules on staging', 'Dev Team', 'Before each phase', 'Not Started', 'No chains, no loops'],
    ['Redirects', 'Backup server config before each phase', 'Dev Team', 'Before each phase', 'Not Started', 'For rollback if needed'],
    ['Content', 'Verify /contact/ page exists (8 backlinks)', 'Content Team', 'Before Phase 2', 'Not Started', 'CRITICAL - missing from original'],
    ['Content', 'Verify /faq/ page exists', 'Content Team', 'Before Phase 2', 'Not Started', '2 backlinks to FAQ'],
    ['Content', 'Verify all city pages exist on new site', 'Content Team', 'Before Phase 3', 'Not Started', '30 city/region pages'],
]

for row_idx, row_data in enumerate(checklist_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        style_data_cell(ws1, row_idx, col_idx, 'center' if col_idx in [3, 4, 5] else 'left')
        if col_idx == 1:
            cell.fill = section_fill
        if col_idx == 5:
            if value == 'Completed':
                cell.fill = success_fill
            elif value == 'In Progress':
                cell.fill = warning_fill
            elif value == 'Not Started':
                cell.fill = gray_fill

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 50
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 35

# ============ SHEET 2: Phase Timeline ============
ws2 = wb.create_sheet('Phase Timeline')

ws2['A1'] = 'Migration Phase Timeline'
ws2['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws2.merge_cells('A1:G1')

headers2 = ['Phase', 'Date', 'Action', 'URLs', 'Owner', 'Priority', 'Status']
for col, header in enumerate(headers2, 1):
    ws2.cell(row=3, column=col, value=header)
style_header_row(ws2, 3, 7)

timeline_data = [
    ['Phase 1', '09-02-2026', 'Legal pages (privacy, cookie, disclaimer)', '3', 'Ronald', 'High', 'Pending'],
    ['Phase 1', '09-02-2026', 'Contact & service pages', '2', 'Ronald', 'High', 'Pending'],
    ['Phase 1', '09-02-2026', 'News articles (VvE related)', '13', 'Ronald', 'Low', 'Pending'],
    ['Phase 1', '10-02-2026', 'Monitor Phase 1 - check GSC for errors', '-', 'SEO Team', 'High', 'Pending'],
    ['Phase 2', '16-02-2026', 'VvE Homepage (43 backlinks - CRITICAL)', '1', 'Brightlot', 'Critical', 'Pending'],
    ['Phase 2', '16-02-2026', 'Contact form page (8 backlinks)', '1', 'Ronald', 'Critical', 'Pending'],
    ['Phase 2', '16-02-2026', 'Package pages (Premium, Plus, Start, etc.)', '7', 'Ronald', 'Critical', 'Pending'],
    ['Phase 2', '16-02-2026', 'Offerte/quote pages', '7', 'Ronald', 'Critical', 'Pending'],
    ['Phase 2', '16-02-2026', 'Customer service pages', '4', 'Ronald', 'High', 'Pending'],
    ['Phase 2', '16-02-2026', 'FAQ pages', '3', 'Ronald', 'Medium', 'Pending'],
    ['Phase 2', '16-02-2026', 'Team/about pages', '2', 'Ronald', 'Medium', 'Pending'],
    ['Phase 2', '17-02-2026', 'Request indexing in GSC for key pages', '-', 'SEO Team', 'High', 'Pending'],
    ['Phase 2', '17-02-2026', 'Monitor Phase 2 - check for 404s', '-', 'SEO Team', 'High', 'Pending'],
    ['Phase 3', '23-02-2026', 'City pages (Amsterdam, Rotterdam, etc.)', '30', 'Ronald', 'High', 'Pending'],
    ['Phase 3', '23-02-2026', 'Region pages (Noord-West, Zuid-Oost, etc.)', '8', 'Ronald', 'High', 'Pending'],
    ['Phase 3', '23-02-2026', 'Sustainability/ESG pages', '2', 'Ronald', 'Medium', 'Pending'],
    ['Phase 3', '23-02-2026', 'Newbuild VvE pages', '2', 'Ronald', 'Medium', 'Pending'],
    ['Phase 3', '24-02-2026', 'Update mvgm.com internal links', '-', 'Brightlot', 'High', 'Pending'],
    ['Phase 3', '24-02-2026', 'Update footer/navigation links', '-', 'Brightlot', 'Medium', 'Pending'],
    ['Post', '01-03-2026', 'Week 1 performance review', '-', 'SEO Team', 'High', 'Pending'],
    ['Post', '15-03-2026', 'Month 1 performance review', '-', 'SEO Team', 'Medium', 'Pending'],
    ['Post', '01-06-2026', 'Full recovery assessment (3 months)', '-', 'SEO Team', 'Medium', 'Pending'],
]

for row_idx, row_data in enumerate(timeline_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        style_data_cell(ws2, row_idx, col_idx, 'center' if col_idx in [1, 2, 4, 5, 6, 7] else 'left')
        if col_idx == 6:
            if value == 'Critical':
                cell.fill = critical_fill
                cell.font = critical_font
            elif value == 'High':
                cell.fill = high_fill
            elif value == 'Medium':
                cell.fill = medium_fill

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12

# ============ SHEET 3: Link Equity Rules ============
ws3 = wb.create_sheet('Link Equity Rules')

ws3['A1'] = 'Link Equity Preservation Rules'
ws3['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws3.merge_cells('A1:D1')

headers3 = ['Rule', 'Do', 'Do Not', 'Why It Matters']
for col, header in enumerate(headers3, 1):
    ws3.cell(row=3, column=col, value=header)
style_header_row(ws3, 3, 4)

rules_data = [
    ['Redirect Type', 'Always use 301 (permanent)', 'Never use 302 (temporary)', '301 passes 90-99% link equity; 302 passes minimal'],
    ['Redirect Chains', 'Direct redirect: A to C', 'Chain redirect: A to B to C', 'Each hop loses ~10% equity'],
    ['Content Match', 'Match topic: /vve-premium/ to /pakket/premium/', 'Generic: /vve-premium/ to /', 'Topical relevance transfers ranking signals'],
    ['URL Structure', 'Preserve semantics: /amsterdam/ to /amsterdam/', 'Change meaning: /amsterdam/ to /cities/', 'Google recognizes semantic similarity'],
    ['Redirect Duration', 'Keep redirects active 12+ months', 'Remove after a few weeks', 'Google needs time to process all redirects'],
    ['Trailing Slashes', 'Handle both: /page and /page/', 'Only redirect one variant', 'Users/links may use either format'],
    ['Query Parameters', 'Map specific params to new URLs', 'Ignore query string redirects', 'Preserves conversion tracking links'],
    ['Canonical Tags', 'Point to new domain only', 'Mix old/new domain canonicals', 'Conflicting signals confuse Google'],
    ['Internal Links', 'Update to point directly to new URLs', 'Rely only on redirects', 'Direct links pass more equity than redirects'],
    ['HTTPS', 'Redirect to HTTPS on new domain', 'Allow HTTP access', 'Security signals affect rankings'],
]

for row_idx, row_data in enumerate(rules_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        style_data_cell(ws3, row_idx, col_idx)
        if col_idx == 2:
            cell.fill = success_fill
        elif col_idx == 3:
            cell.fill = critical_fill
            cell.font = Font(color='FFFFFF')

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 40
ws3.column_dimensions['D'].width = 45

# ============ SHEET 4: Post-Migration Monitoring ============
ws4 = wb.create_sheet('Monitoring Schedule')

ws4['A1'] = 'Post-Migration Monitoring Schedule'
ws4['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws4.merge_cells('A1:F1')

headers4 = ['Period', 'Check', 'Tool', 'Target', 'Actual', 'Status']
for col, header in enumerate(headers4, 1):
    ws4.cell(row=3, column=col, value=header)
style_header_row(ws4, 3, 6)

monitoring_data = [
    ['Week 1 (Daily)', 'Crawl errors', 'GSC - Coverage', '0 new errors', '', 'Pending'],
    ['Week 1 (Daily)', '404 errors', 'GSC - Coverage - Excluded', '<5 new 404s', '', 'Pending'],
    ['Week 1 (Daily)', 'New pages indexed', 'site:mvgm-vvebeheer.nl', 'Pages appearing', '', 'Pending'],
    ['Week 1 (Daily)', 'Traffic drop', 'Google Analytics', '<20% drop (normal)', '', 'Pending'],
    ['Week 1 (Daily)', 'Server errors', 'Server logs', '0 500 errors', '', 'Pending'],
    ['Week 2-4 (Weekly)', 'Ranking positions', 'Semrush/Ahrefs', 'Stabilizing', '', 'Pending'],
    ['Week 2-4 (Weekly)', 'Organic traffic', 'Google Analytics', 'Recovering', '', 'Pending'],
    ['Week 2-4 (Weekly)', 'Backlinks appearing', 'GSC - Links', 'New domain showing', '', 'Pending'],
    ['Week 2-4 (Weekly)', 'Redirect errors', 'Screaming Frog', '0 chain/loops', '', 'Pending'],
    ['Month 2-3', 'Full traffic recovery', 'Google Analytics', '>=90% of pre-migration', '', 'Pending'],
    ['Month 2-3', 'Ranking recovery', 'Rank tracker', 'Same or improved', '', 'Pending'],
    ['Month 2-3', 'Old URLs de-indexed', 'site:mvgm.com vve-beheer', 'Decreasing count', '', 'Pending'],
    ['Month 2-3', 'Conversion recovery', 'Google Analytics', '>=90% of pre-migration', '', 'Pending'],
]

for row_idx, row_data in enumerate(monitoring_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        style_data_cell(ws4, row_idx, col_idx, 'center' if col_idx in [1, 4, 5, 6] else 'left')
        if col_idx == 1:
            cell.fill = section_fill

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 25
ws4.column_dimensions['E'].width = 20
ws4.column_dimensions['F'].width = 12

# ============ SHEET 5: Common Mistakes ============
ws5 = wb.create_sheet('Common Mistakes')

ws5['A1'] = 'Common Migration Mistakes to Avoid'
ws5['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws5.merge_cells('A1:D1')

headers5 = ['Mistake', 'Problem', 'Solution', 'How to Check']
for col, header in enumerate(headers5, 1):
    ws5.cell(row=3, column=col, value=header)
style_header_row(ws5, 3, 4)

mistakes_data = [
    ['Soft 404s', 'Page returns 200 but shows error content', 'Ensure all destinations have real content', 'Manual check + Screaming Frog'],
    ['Wrong Canonical', 'New page canonical points to old domain', 'Update all canonicals to mvgm-vvebeheer.nl', 'View source, check link rel canonical'],
    ['Mixed HTTP/HTTPS', 'Internal links use http instead of https', 'Audit all internal links, ensure HTTPS', 'Screaming Frog protocol report'],
    ['Blocking Googlebot', 'robots.txt blocks crawling or noindex tags', 'Check robots.txt and meta robots before go-live', 'GSC URL Inspection tool'],
    ['Forgetting Internal Links', 'mvgm.com still links to old VvE URLs', 'Update internal links to new domain directly', 'Crawl mvgm.com, check for old URLs'],
    ['Using 302 Instead of 301', '302 passes minimal link equity', 'Always use 301 for permanent moves', 'curl -I [url] to check response code'],
    ['Redirect Chains', 'A to B to C loses equity at each hop', 'All redirects go directly to final URL', 'Redirect checker tool'],
    ['Not Monitoring', 'Issues go unnoticed for weeks', 'Set calendar reminders for checks', 'GSC email alerts for crawl errors'],
    ['Removing Redirects Too Soon', 'Old backlinks stop working', 'Keep redirects active 12+ months', 'Do not remove from server config'],
    ['Missing Query Params', 'URLs with ?parameters return 404', 'Map all query string variations', 'Check GSC for 404s with params'],
    ['No Rollback Plan', 'Cannot recover if something breaks', 'Backup config before each phase', 'Test rollback on staging first'],
    ['Skipping Benchmarks', 'Cannot measure success/failure', 'Document traffic and rankings before', 'Export reports before Phase 1'],
]

for row_idx, row_data in enumerate(mistakes_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        style_data_cell(ws5, row_idx, col_idx)
        if col_idx == 1:
            cell.fill = critical_fill
            cell.font = critical_font
        elif col_idx == 3:
            cell.fill = success_fill

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 40
ws5.column_dimensions['C'].width = 45
ws5.column_dimensions['D'].width = 35

# ============ SHEET 6: Success Metrics ============
ws6 = wb.create_sheet('Success Metrics')

ws6['A1'] = 'Migration Success Metrics'
ws6['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws6.merge_cells('A1:E1')

headers6 = ['Metric', 'Acceptable', 'Good', 'Excellent', 'Your Result']
for col, header in enumerate(headers6, 1):
    ws6.cell(row=3, column=col, value=header)
style_header_row(ws6, 3, 5)

metrics_data = [
    ['Traffic Recovery (3 months)', '80%', '90%', '100%+', ''],
    ['Ranking Recovery (3 months)', '-5 positions', 'Same', 'Improved', ''],
    ['Crawl Errors', '<10', '<5', '0', ''],
    ['404 Errors', '<20', '<10', '0', ''],
    ['Redirect Chains', '<5', '1-2', '0', ''],
    ['Indexing Speed (weeks)', '4-6', '2-3', '1', ''],
    ['Conversion Recovery', '80%', '90%', '100%+', ''],
    ['Backlink Transfer', '80%', '90%', '95%+', ''],
]

light_green = PatternFill('solid', fgColor='90EE90')

for row_idx, row_data in enumerate(metrics_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=value)
        style_data_cell(ws6, row_idx, col_idx, 'center')
        if col_idx == 2:
            cell.fill = warning_fill
        elif col_idx == 3:
            cell.fill = light_green
        elif col_idx == 4:
            cell.fill = success_fill

ws6.column_dimensions['A'].width = 30
ws6.column_dimensions['B'].width = 15
ws6.column_dimensions['C'].width = 15
ws6.column_dimensions['D'].width = 15
ws6.column_dimensions['E'].width = 15

# Save workbook
wb.save('09-seo-migration-best-practices.xlsx')
print('Excel file created successfully!')
